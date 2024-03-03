import logging
import random
import threading
import time
from datetime import datetime, timedelta, timezone
from logging.handlers import RotatingFileHandler
from typing import List
import uuid
import pytz
import telegram
from fastapi import (BackgroundTasks, Depends, FastAPI, File, Header,
                     HTTPException, Path, Query, UploadFile)
from fastapi.background import BackgroundTasks
from fastapi.responses import JSONResponse, PlainTextResponse
from openpyxl import load_workbook
from sqlalchemy import (Boolean, Column, DateTime, Integer, String,
                        create_engine, desc, inspect,distinct, true,Sequence)
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.sql.expression import false
from telegram import (InlineKeyboardButton, InlineKeyboardMarkup,
                      ReplyKeyboardMarkup, ReplyKeyboardRemove, Update)
from telegram.ext import (CallbackContext, CallbackQueryHandler,
                          CommandHandler, ContextTypes, ConversationHandler,
                          Filters, MessageHandler, Updater, filters)

# configure the log format
formatter  = logging.Formatter('%(asctime)s - %(message)s')


handler = RotatingFileHandler('app.log', maxBytes=1024*1024*10, backupCount=5)
handler.setFormatter(formatter)
# set the logging level to INFO
logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
logging.getLogger('sqlalchemy').setLevel(logging.ERROR)

# add the handler to the logger
logger = logging.getLogger(__name__)
logger.addHandler(handler)

description = """

## Isnad Tasks - Util API <img src=\'https://flagcdn.com/24x18/ps.png\'> 🔻🔻🔻


Isnad tasks - is a powerful tool designed to streamline Isnad tasks.

The following APIs provide various services for managing tasks, handling target user IDs, and reading file contents.

**Key Features:**

- **Upload Tasks:**
    Upload the main Twitter tasks, used by Isnad team. 
    The uploaded file should follow the format: `TASK_URL|TASK_TARGET_TYPE`, with each record on a new line.

- **Upload Target User IDs:**
    Maintain a list of target Twitter users. 
    Each line in the uploaded file should represent a single Twitter user ID.

**Authentication:**
    
Access to these services is protected by an API key mechanism. Users must provide a valid API key in the request header for authentication.

**How to Use:**
 
- To upload target users: Use the `/upload-target-accounts/` endpoint, ensuring the provided API key is valid.

- To upload excel file of Isnad tasks: Use the `/upload-isnad-tasks/` endpoint, providing an excel sheet of the accounts details and ensuring the provided API key is valid.

- To Check Target Account Details: Access the `/get-target-account-details/` specifying the account name and providing the API key for authentication.

- To display logs: Access the `/logs/` endpoint.
 

**Obtaining an API Key:**
    
For users requiring an API key, please contact `M Mansour` for assistance.

"""

app = FastAPI(title="Isnad Tasks Bot",
              description=description,
              summary="Isnad Tasks - Util API.",
              version="0.0.1", swagger_ui_parameters={"defaultModelsExpandDepth": -1}
              )

# SQLite database setup
DATABASE_URL = "sqlite:///./isnad.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
Base = declarative_base()


class TargetAccount(Base):
    __tablename__ = "target_accounts"

    id = Column(Integer, primary_key=True, index=True)
    account_name = Column(String, index=True)
    account_id = Column(String, index=True)
    account_link = Column(String)
    account_status = Column(String)
    account_category = Column(String,index=True)
    account_type = Column(String,index=True)
    publishing_level = Column(String,index=True)
    access_level = Column(String,index=True)
    is_used = Column(Boolean, default=false(),index=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class IsnadTasks(Base):
    __tablename__ = "isnad_tasks"

    id = Column(Integer, primary_key=True, index=True)
    task_url = Column(String, index=True)
    task_target_type = Column(String, index=True)
    is_used = Column(Boolean, default=false(),index=True)
    batch_id = Column(Integer, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)


# Define a Task class to represent the tasks table in the database
class Task(Base):
    __tablename__ = 'tasks1'

    id = Column(Integer, primary_key=True)
    name = Column(String)
    used = Column(Boolean, default=False)


Base.metadata.create_all(bind=engine)

# Define a dictionary to track tasks used by each user
user_tasks = {}
user_sessions = {}


# Function to add dummy tasks data for testing
def add_dummy_tasks():
    Session = sessionmaker(bind=engine)
    session = Session()
    workbook = load_workbook('FINAL_IDs.xlsx')
    sheet = workbook.active

    # Map column names to indices
    header = sheet[1]
    column_indices = {header[i].value: i for i in range(len(header))}

    # Extract data from the Excel sheet and add or update it in the database
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print('row[column_indices["ACCOUNT_NAME"]]:',row[column_indices["ACCOUNT_NAME"]])
        account_data = {
            "account_name": row[column_indices["ACCOUNT_NAME"]],
            "account_id": row[column_indices["ACCOUNT_ID"]],
            "account_link": row[column_indices["ACCOUNT_LINK"]],
            "account_status": row[column_indices["ACCOUNT_STATUS"]],
            "account_category": row[column_indices["ACCOUNT_CATEGORY"]],
            "account_type": row[column_indices["ACCOUNT_TYPE"]],
            "publishing_level": row[column_indices["PUBLISHING_LEVEL"]],
            "access_level": row[column_indices["ACCESS_LEVEL"]],
        }

        # Check if the record exists based on unique identifier (account_id)
        existing_record = session.query(TargetAccount).filter(TargetAccount.account_name == account_data["account_name"]).first()

        if existing_record:
            # Update existing record with new values
            for key, value in account_data.items():
                setattr(existing_record, key, value)
            else:
                # Create a new record and add it to the database
                db_account = TargetAccount(**account_data)
                session.add(db_account)

    session.commit()       
   

def get_db():
    # Step 5: Insert data into the database
    Session = sessionmaker(bind=engine)
    session = Session()
    # db = session
    try:
        yield session
    finally:
        session.close()


def is_database_empty():
    """
    Check if the accounts table is empty.

    - **db**: Database session.

    Returns a boolean indicating whether the accounts table is empty.
    """
    Session = sessionmaker(bind=engine)
    session = Session()
    return not session.query(TargetAccount).first()


def check_database_status():
    """
    Check the status of the database and accounts table.

    - **db**: Database session.

    Returns a tuple (is_database_exist, is_table_exist, is_table_empty).
    """
    inspector = inspect(engine)
    is_database_exist = inspector.get_schema_names()

    if is_database_exist:
        is_table_exist = inspector.get_table_names()

        if is_table_exist and "target_accounts" in is_table_exist:
            is_table_empty = is_database_empty()
            return True, True, is_table_empty

    return False, False, False

# Define a function to generate unique batch IDs
def generate_batch_id():
    Session = sessionmaker(bind=engine)
    session = Session()
    # Get the maximum batch ID from the database and increment it
    max_batch_id = session.query(IsnadTasks.batch_id).order_by(IsnadTasks.batch_id.desc()).first()
    if max_batch_id is None:
        return 1
    else:
        return max_batch_id[0] + 1


API_KEY_ADMIN = "iSLgvYQMFbExJGIVpJHEOEHnYxyzT4Fcr5xfSVG2Sn0q5FcrylK72Pgs3ctg0Cyp"


# Define a dictionary to store the mapping of userid to api_key
user_api_key_map = {
    "user1": "Hw1MXuWmKwsG4UXlRITVvS3vkKd5xvkiKD2Z9lXPvXZ5tuUEsTGAfqT8m8AnNGuo",
    "user2": "ERGZdjZqumtfZccYmpYwIlAO83RrqATioi6OUXQI8iiVZtG3xiKBfGgPjqgMwdvw",
    "user3": "WbpClFZ2HnsNgbYBwsoYeVFqUGYu64a71Thj7qHA9xE7ca8zjKFw1rOQzohwVOKX",
    "user4": "wRkBnrIMx20TPYRduKsq3SXfc8WkXh0Pj3H0hGYGJBT7qoXXxYMzTMk4JUqkyMsl",
    "user5": "AfIZUKMWVNo0KDnMdinHqaFIZnDgEWzDBw2PgubmffcQzUj9Lh5WaTz3ilzFx8Dp",
    "user6": "XQ5ihKJl2GqUvMM8O0Fs06UmZy6d0EeF4u3QLAMVbppzJETTul90PhQh7vI9oC4R",
    "user7": "fEpvjfO0oZr4fb3ncjQyAYdOc3DdkiCEhlKfBiNa8biHRHTly2duw20C44QZHBCf",
    "user8": "fKrFNeOwapEe7XrIiTRl9ufMbmxEaNGazYpemjb2VVkS8Z40fYVtgQMC46A26K7o",
    "user9": "wJkclhGS7ROCf13YncA69meOp7sdK5iAp9ofMYxbAUc9Gm7fFJ94Xj8EEnTqIEDq",
    "user10": "RpnNbdW9sN8zVddnuHrsFG0I2VMzY2VBI5tnrHhYqVrXiovvNsqHBNkpwjSwyyJf",
    "user11": "SedkImeAadVNu6TloPajo4ekQohXe5yWAi7aEb7aeeHVOeYqvI464Uj9cVXhjoVF",
    "user12": "onHGZ9U57aYeRWUiwoIpdRK4DVhCybGCQLO6xHXtW4SimPtas3j8f4K6OPpTVCEh",
    "admin": API_KEY_ADMIN
    # Add more users and their corresponding api keys
}


# Dependency to get the userid based on the provided api_key
async def get_api_key(api_key: str = Header(..., description="API key for authentication")):
    """
    Get User ID

    Retrieves the userid based on the provided API key.

    - **api_key**: API key for authentication.

    Returns the corresponding userid.
    """
    for userid, key in user_api_key_map.items():
        if key == api_key:
            return userid
    raise HTTPException(
        status_code=401,
        detail="Invalid API key",
        headers={"WWW-Authenticate": "Bearer"},
    )

# Dependency to check the admin API key


def get_admin_api_key(api_key: str = Header(..., description="Admin API key for authentication")):
    if api_key != API_KEY_ADMIN:
        raise HTTPException(
            status_code=401,
            detail="Invalid admin API key",
        )
    return api_key


@app.get("/", include_in_schema=False)
def read_root():
    return {"Isnad ": "Tasks🔻🔻🔻"}

@app.on_event("startup")
async def startup_event():
    print('Isnad Tasks Server started ---- at: ', datetime.now())
    try:
        main()
        pass
    except telegram.error.Conflict as e:
        print(f"Telegram Conflict Error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


# Define a function to check if the file is a text file
def is_text_file(file: UploadFile):
    if not file.content_type.startswith("text/"):
        raise HTTPException(
            status_code=400,
            detail="Only text files are allowed",
        )

# Dependency to check if the file is an Excel file
def is_excel_file(file: UploadFile):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Only text files are allowed",
        )
    return file


# Endpoint to upload Excel file and add data to DB
@app.post("/upload-target-accounts/")
async def upload_target_accounts(
    api_key: str = Depends(get_api_key),
    file: UploadFile = Depends(is_excel_file),
    db: Session = Depends(get_db)
):
    """
    Upload Excel File of Twitter target accounts, and Add or Update Data in Database

    Allows the application to upload an Excel file, extract data, and add or update it in the database.

    - **file**: Upload an Excel file.
    - **db**: Database session.

    Returns a confirmation message.
    """
    try:
        # Load Excel file and extract data
        workbook = load_workbook(file.file)
        sheet = workbook.active

        # Map column names to indices
        header = sheet[1]
        column_indices = {header[i].value: i for i in range(len(header))}

        # Extract data from the Excel sheet and add or update it in the database
        for row in sheet.iter_rows(min_row=2, values_only=True):
            account_data = {
                "account_name": row[column_indices["ACCOUNT_NAME"]],
                "account_id": row[column_indices["ACCOUNT_ID"]],
                "account_link": row[column_indices["ACCOUNT_LINK"]],
                "account_status": row[column_indices["ACCOUNT_STATUS"]],
                "account_category": row[column_indices["ACCOUNT_CATEGORY"]],
                "account_type": row[column_indices["ACCOUNT_TYPE"]],
                "publishing_level": row[column_indices["PUBLISHING_LEVEL"]],
                "access_level": row[column_indices["ACCESS_LEVEL"]],
            }

            # Check if the record exists based on unique identifier (account_id)
            existing_record = db.query(TargetAccount).filter(TargetAccount.account_id == account_data["account_id"]).first()

            if existing_record:
                # Update existing record with new values
                for key, value in account_data.items():
                    setattr(existing_record, key, value)
            else:
                # Create a new record and add it to the database
                db_account = TargetAccount(**account_data)
                db.add(db_account)

        db.commit()
        logger.info('Request from UserID: ' +
                    api_key+' - Target Accounts, Data added or updated in the database successfully')
        return JSONResponse(content={"message": "Target Accounts, Data added or updated in the database successfully."}, status_code=200)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Target Accounts Excel file: {str(e)}")


# Endpoint to upload Excel file and add data to DB
@app.post("/upload-isnad-tasks/")
async def upload_isnad_tasks(
    api_key: str = Depends(get_api_key),
    file: UploadFile = Depends(is_excel_file),
    db: Session = Depends(get_db)
):
    """
    Upload Excel File of Twitter Isnad accounts, and Add or Update Data in Database

    Allows the application to upload an Excel file, extract data, and add or update it in the database.

    - **file**: Upload an Excel file.
    - **db**: Database session.

    Returns a confirmation message.
    """
    try:
        # Load Excel file and extract data
        workbook = load_workbook(file.file)
        sheet = workbook.active
        # Delete all current tasks
        db.query(IsnadTasks).delete()
        # Map column names to indices
        header = sheet[1]
        column_indices = {header[i].value: i for i in range(len(header))}
        # batch_id = generate_batch_id()
        batch_id = uuid.uuid4().hex.upper()[0:6]
        # Extract data from the Excel sheet and add or update it in the database
        for row in sheet.iter_rows(min_row=2, values_only=True):
            account_data = {
                "task_url": row[column_indices["TASK_URL"]],
                "task_target_type": row[column_indices["TASK_TARGET_TYPE"]],
                "batch_id": batch_id
                }
            db_account = IsnadTasks(**account_data)
            db.add(db_account)

        db.commit()
        logger.info('Request from UserID: ' +
                    api_key+' - Isnad tasks, Data added or updated in the database successfully')
        return JSONResponse(content={"message": "Isnad tasks , Data added or updated in the database successfully."}, status_code=200)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Isnad Tasks Excel file: {str(e)}")


# Endpoint to get account details by ACCOUNT_NAME
@app.get("/get-target-account-details/")
async def get_account(
        api_key: str = Depends(get_api_key),
        account_name: str = Query(..., title="Account Id", description="The ACCOUNT_NAME to retrieve details for."), db: Session = Depends(get_db)):
    """
    Get Target Account Details by ACCOUNT_NAME

    Allows the application to retrieve account details from the database based on the provided ACCOUNT_NAME.

    - **account_name**: The ACCOUNT_NAME to retrieve details for.
    - **db**: Database session.

    Returns account details.
    """
    
    account = db.query(TargetAccount).filter(TargetAccount.account_name == account_name).first()

    if not account:
        raise HTTPException(
            status_code=404,
            detail="Account not found",
        )

    logger.info('Request from UserID: '+api_key +
                ' - Search for ACCOUNT_NAME: '+account_name+' .')

    return {
        "account_name": account.account_name,
        "account_id": account.account_id,
        "account_link": account.account_link,
        "account_status": account.account_status,
        "account_category": account.account_category,
        "account_type": account.account_type,
        "publishing_level": account.publishing_level,
        "access_level": account.access_level,
        "is_used": account.is_used,
        "created_at": account.created_at
    }


# Log viewer
@app.get("/logs", response_class=PlainTextResponse)
async def read_logs(api_key: str = Depends(get_api_key)):
    """
    View Application Logs

    Displays the application logs.

    Returns logs.
    """
    try:
        with open("app.log", "r", encoding="utf-8") as file:
            logs = file.readlines()

        # Reverse the order of logs to get the most recent entries first
        logs.reverse()

        content = "".join(logs)
        return content
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"File app.log not found.",
        )

# Define the private group ID
ISNAD_GROUP_ID = -1002038467827  # Replace with your private group ID

# Define your welcome message and options
welcome_message = "أهلا بك في بوت مهمات *إسناد.*\n\n لبدأ المهمات, برجاء الضغط علي الإختيار التالي  . .\n\n"

def has_completed_batch(user_id: int, batch_id: int, db: Session) -> bool:
    return db.query(IsnadTasks).filter(IsnadTasks.batch_id == batch_id, IsnadTasks.is_used == true()).count() == db.query(IsnadTasks).filter(IsnadTasks.batch_id == batch_id).count()

# Define a function to handle the /start command
def start(update: Update, context: CallbackContext) -> None:
    """Send a welcome message with options when the command /start is issued."""
    # Get the user ID of the user who triggered the command
    user_id = update.effective_user.id

    # Check if the user is a member of the private group
    is_member = context.bot.get_chat_member(ISNAD_GROUP_ID, user_id)

    if is_member.status != 'left':
        # If the user is a member of the private group, send a welcome message
        # update.message.reply_text("Welcome! You are authorized to use this bot.")
        keyboard = [
        [InlineKeyboardButton("💥 مهمة اليوم الجديدة", callback_data='option1')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_chat.id,text=welcome_message, reply_markup=reply_markup,  parse_mode= 'Markdown')
    else:
        # If the user is not a member of the private group, inform them that they need to join the group to use the bot
        update.message.reply_text("Sorry, you are not authorized to use this bot.")

    
# Function to fetch the next task for a user
def get_next_task(user_id):
    Session = sessionmaker(bind=engine)
    session = Session()

    # Get the current batch_id from the IsnadTasks table
    current_batch_ids = session.query(distinct(IsnadTasks.batch_id)).all()
    # Extract the batch_id from the result
    current_batch_id = current_batch_ids[0][0] if current_batch_ids else None
    # Fetch all tasks
    all_tasks = session.query(IsnadTasks).all()

    # Fetch tasks used by the user
    user_used_tasks = user_tasks.get(user_id, [])
    # Initialize next_task outside the conditional block
    next_task = None

    # Check if user has any used tasks
    if user_used_tasks:
        # Get the batch_id of the last task used by the user
        last_batch_id = user_used_tasks[-1][1]
        # Extract list of used task IDs from user_used_tasks
        used_task_ids = [task_id for task_id, _ in user_used_tasks]
        # Determine unused tasks not used by the user
        available_tasks = [task for task in all_tasks if not task.is_used and task.id not in used_task_ids and task.batch_id == last_batch_id]

        # Check if there are any available tasks in the same batch
        if available_tasks:
            next_task = random.choice(available_tasks)

    # If no next_task is found for the same batch, find a new one
    if not next_task:
        if user_used_tasks and current_batch_id==last_batch_id:
            return None
        else:
            # Determine unused tasks not used by the user
            available_tasks = [task for task in all_tasks if not task.is_used]

            if available_tasks:
                # Select a random task from available tasks
                next_task = random.choice(available_tasks)
            else:
                # Reset used flag for all tasks
                session.query(IsnadTasks).update({IsnadTasks.is_used: False})
                session.commit()
                # Select a random task
                next_task = random.choice(all_tasks)

            # Mark the task as used
            next_task.is_used = True
            session.commit()

    # Check if the user has used tasks from a previous batch
    if user_used_tasks and user_used_tasks[-1][1] != next_task.batch_id:
        # Clear the dictionary for the user and insert a new one for the new batch id
        user_tasks[user_id] = []

    # Update user_tasks dictionary with the current task
    user_tasks.setdefault(user_id, []).append((next_task.id, next_task.batch_id))

    return next_task



# Define a function to handle button clicks
def button_click(update: Update, context: CallbackContext) -> None:
    """Respond to button clicks."""
    query = update.callback_query
    query.answer()
    option = query.data
    Session = sessionmaker(bind=engine)
    session = Session()

    if option == 'option1':
        next_task = get_next_task(update.callback_query.from_user.id)
        if next_task:

            user_sessions[update.callback_query.from_user.id] = {"task_target_type": next_task.task_target_type}
            print(f"Next task for user {update.callback_query.from_user.id}: {next_task.id}")
            query.message.reply_text(text=f"<b>انسخ منشور جديد:</b>: \n\n {next_task.task_url}",  
                                parse_mode= 'HTML',disable_web_page_preview=True)

            target_type = next_task.task_target_type
            if target_type and int(target_type) < 1:
                target_accounts = session.query(TargetAccount).filter(
                TargetAccount.account_type == next_task.task_target_type, TargetAccount.is_used == false()).order_by(
                    TargetAccount.publishing_level.asc(), TargetAccount.access_level.asc()).limit(4).all()
            else:
                target_accounts = session.query(TargetAccount).filter(
                    TargetAccount.account_type == next_task.task_target_type, TargetAccount.is_used == false()).order_by(
                        TargetAccount.publishing_level.asc(), TargetAccount.access_level.asc()).limit(4).all()
                # Check if the query returned no unused accounts
            if not target_accounts:
                # Reset is_used for all accounts to False and retry the query
                if target_type and int(target_type) < 1:
                    session.query(TargetAccount).filter(
                        TargetAccount.account_type == next_task.task_target_type).update({TargetAccount.is_used: false()})
                else:
                    session.query(TargetAccount).update({TargetAccount.is_used: false()})
                session.commit()

                if target_type and int(target_type) < 1:
                    target_accounts = session.query(TargetAccount).filter(
                        TargetAccount.account_type == next_task.task_target_type,
                        TargetAccount.is_used == false()).order_by(TargetAccount.publishing_level.asc(),
                                                            TargetAccount.access_level.asc()).limit(4).all()
                else:
                    target_accounts = session.query(TargetAccount).filter(TargetAccount.is_used == false()).order_by(TargetAccount.publishing_level.asc(),
                                                            TargetAccount.access_level.asc()).limit(4).all()
            if target_accounts:
                query.message.reply_text(text= "<b>الحسابات المستهدفة:</b>",  
                                                parse_mode= 'HTML',disable_web_page_preview=True)
                for target_account in target_accounts:
                    if target_account.account_id:
                        # Check the stop flag before processing each account
                        query.message.reply_text(text=f"{target_account.account_link}",  
                                                parse_mode= 'HTML',disable_web_page_preview=True)
                        target_account.is_used = True
                        # Commit changes to the database
                    session.commit()
                    
                keyboard = [
                [InlineKeyboardButton("🔄 مهمة اليوم الجديدة",  callback_data='option1')],
                [InlineKeyboardButton("🔻اكونتات مستهدفة جديدة🔻", callback_data='option2')]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)

                query.message.reply_text(' يرجي إختيار أحد الخيارات التالية: ', reply_markup=reply_markup) 
        else:
            keyboard = [
                [InlineKeyboardButton("🔄 مهمة اليوم الجديدة",  callback_data='option1')]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            query.message.reply_text('🔻انتهت مهمة اليوم، ربنا يسدد رمينا ',reply_markup=reply_markup) 
    if option == 'option2':
        task_target_type = user_sessions.get(update.callback_query.from_user.id, {}).get("task_target_type")
        if task_target_type:
            target_accounts = session.query(TargetAccount).filter(
                TargetAccount.account_type == task_target_type, TargetAccount.is_used == false()).order_by(
                    TargetAccount.publishing_level.asc(), TargetAccount.access_level.asc()).limit(4).all()
                # Check if the query returned no unused accounts
            if not target_accounts:
                # Reset is_used for all accounts to False and retry the query
                session.query(TargetAccount).filter(
                    TargetAccount.account_type == task_target_type).update({TargetAccount.is_used: false()})
                session.commit()

                target_accounts = session.query(TargetAccount).filter(
                    TargetAccount.account_type == task_target_type,
                    TargetAccount.is_used == false()).order_by(TargetAccount.publishing_level.asc(),
                                                            TargetAccount.access_level.asc()).limit(4).all()
            if target_accounts:
                query.message.reply_text(text= "<b>الحسابات المستهدفة:</b>",  
                                            parse_mode= 'HTML',disable_web_page_preview=True)
                for target_account in target_accounts:
                    if target_account.account_id:
                        # Check the stop flag before processing each account
                        query.message.reply_text(text=f"{target_account.account_link}",  
                                                parse_mode= 'HTML',disable_web_page_preview=True)
                        target_account.is_used = True
                        # Commit changes to the database
                    session.commit()
                    
                keyboard = [
                [InlineKeyboardButton("🔄 مهمة اليوم الجديدة",  callback_data='option1')],
                [InlineKeyboardButton("🔻اكونتات مستهدفة جديدة🔻", callback_data='option2')]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                query.message.reply_text(' يرجي إختيار أحد الخيارات التالية: ', reply_markup=reply_markup) 
            else:
                query.message.reply_text("You need to select a task first.")
                   

def main() -> None:

    # Add dummy tasks data
    # add_dummy_tasks()
    # print("Dummy tasks data added.")
    """Run the bot."""
    # Create the Updater and pass it your bot's token
    updater = Updater("6930798784:AAF0t4shccWVHyoieRDuJfpKHqr-_YlxiPw")

    # Get the dispatcher to register handlers
    dispatcher = updater.dispatcher

    # Register the handlers
    dispatcher.add_handler(CommandHandler("start", start))
    dispatcher.add_handler(CallbackQueryHandler(button_click))
    # Every time the back button is pressed, the main_menu fucntion is triggered and the user sees the previous menu
    dispatcher.add_handler(CallbackQueryHandler(button_click, pattern='back'))

    # Start the Bot
    updater.start_polling()

    # Run the bot until you press Ctrl-C
    # updater.idle()

if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
